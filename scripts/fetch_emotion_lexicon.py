"""One-time setup: download the NRC Emotion Lexicon (EmoLex) and NRC VAD
Lexicon's Romanian translations, and cache a compact, pre-processed form
for `preprocess/emotion.py` (preprocess/objectives.md Phase 6).

License note (important): the NRC lexicons are "free for non-commercial
research and educational purposes" but the terms of use explicitly say
"do not redistribute the data" and require citation. Unlike REDv2/RONEC
(MIT, cached into data/external/ and committed), this data is fetched
fresh into `.cache/lexicons/` (gitignored) and never committed — re-run
this script to (re)populate the cache. See data/external/SOURCES.md for
the full citation.

Source: https://saifmohammad.com/WebPages/AccessResource.htm
(non-commercial research use direct download).
"""

from __future__ import annotations

import io
import json
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path

CACHE_DIR = Path(__file__).resolve().parents[1] / ".cache" / "lexicons"

SUITE_URL = "https://saifmohammad.com/WebDocs/Lexicons/NRC-Suite-of-Sentiment-Emotion-Lexicons.zip"

_PREFIX = "NRC-Suite-of-Sentiment-Emotion-Lexicons/NRC-Sentiment-Emotion-Lexicons"
EMOLEX_PATH = f"{_PREFIX}/NRC-Emotion-Lexicon-v0.92/NRC-Emotion-Lexicon-v0.92-In105Languages-Nov2017Translations.xlsx"
VAD_PATH = f"{_PREFIX}/NRC-VAD-Lexicon/OneFilePerLanguage/Romanian-ro-NRC-VAD-Lexicon.txt"

ROMANIAN_COLUMN = 74  # 0-based, in EMOLEX_PATH's header row
EMOTION_COLUMNS = {
    107: "anger",
    108: "anticipation",
    109: "disgust",
    110: "fear",
    111: "joy",
    112: "sadness",
    113: "surprise",
    114: "trust",
}

CITATION = """\
NRC Emotion Lexicon (EmoLex) and NRC VAD Lexicon, National Research
Council Canada, Dr. Saif M. Mohammad. Non-commercial research/educational
use. Cite:
  Saif Mohammad and Peter Turney. "Crowdsourcing a Word-Emotion Association
  Lexicon." Computational Intelligence, 29(3), 2013.
  Saif Mohammad. "Obtaining Reliable Human Ratings of Valence, Arousal,
  and Dominance for 20,000 English Words." ACL 2018.
See https://saifmohammad.com/WebPages/lexicons.html
"""


def _download_zip() -> bytes:
    # The server rejects urllib's default User-Agent (406 Not Acceptable).
    request = urllib.request.Request(SUITE_URL, headers={"User-Agent": "curl/8.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def _extract_emolex(zip_bytes: bytes) -> dict[str, list[str]]:
    import openpyxl

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        with archive.open(EMOLEX_PATH) as handle:
            workbook = openpyxl.load_workbook(io.BytesIO(handle.read()), read_only=True)

    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(min_row=2, values_only=True)

    merged: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        romanian = row[ROMANIAN_COLUMN]
        if not romanian or romanian.strip().upper() == "NO TRANSLATION":
            continue
        emotions = {name for col, name in EMOTION_COLUMNS.items() if row[col]}
        if not emotions:
            continue
        merged[romanian.strip().lower()] |= emotions

    return {word: sorted(emotions) for word, emotions in merged.items()}


def _extract_vad(zip_bytes: bytes) -> dict[str, dict[str, float]]:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        text = archive.read(VAD_PATH).decode("utf-8")

    lines = text.splitlines()
    header = lines[0].split("\t")
    assert header == ["Word", "Romanian-ro", "Valence", "Arousal", "Dominance"], header

    sums: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
    counts: dict[str, int] = defaultdict(int)
    for line in lines[1:]:
        if not line.strip():
            continue
        _, romanian, valence, arousal, dominance = line.split("\t")
        if romanian.strip().upper() == "NO TRANSLATION":
            continue
        key = romanian.strip().lower()
        sums[key][0] += float(valence)
        sums[key][1] += float(arousal)
        sums[key][2] += float(dominance)
        counts[key] += 1

    return {
        word: {
            "valence": sums[word][0] / counts[word],
            "arousal": sums[word][1] / counts[word],
            "dominance": sums[word][2] / counts[word],
        }
        for word in sums
    }


def main() -> None:
    print("Downloading NRC lexicon suite (~109MB, one-time)...")
    zip_bytes = _download_zip()

    print("Extracting Romanian EmoLex...")
    emolex = _extract_emolex(zip_bytes)

    print("Extracting Romanian VAD...")
    vad = _extract_vad(zip_bytes)

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    (CACHE_DIR / "nrc_emolex_ro.json").write_text(
        json.dumps(emolex, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (CACHE_DIR / "nrc_vad_ro.json").write_text(
        json.dumps(vad, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (CACHE_DIR / "CITATION.txt").write_text(CITATION, encoding="utf-8")

    print(f"Cached {len(emolex)} EmoLex words and {len(vad)} VAD words to {CACHE_DIR}")
    print()
    print(CITATION)


if __name__ == "__main__":
    main()
